import sqlite3
from openpyxl import Workbook

print("=" * 80)
print("DAY 3: EXPORT RFM TO EXCEL")
print("=" * 80)

# Load data from database
conn = sqlite3.connect('data/ecommerce.db')
cursor = conn.cursor()

print("\n1. Reading RFM data from database...")

cursor.execute('''
SELECT 
    c.customer_id,
    c.customer_name,
    c.city,
    MAX(o.order_date) as last_purchase,
    COUNT(o.order_id) as frequency,
    SUM(o.order_amount) as monetary
FROM customers c
LEFT JOIN orders o ON c.customer_id = o.customer_id
GROUP BY c.customer_id, c.customer_name, c.city
ORDER BY monetary DESC
''')

results = cursor.fetchall()

print(f"   {len(results)} customers data retrieved")

# Create Excel workbook
print("\n2. Creating Excel workbook...")

wb = Workbook()
ws = wb.active
ws.title = "RFM Analysis"

# Write headers
headers = ["Customer ID", "Name", "City", "Last Purchase", "Frequency", "Monetary"]
ws.append(headers)

# Write data
print("\n3. Writing data to Excel...")

for row in results:
    ws.append(row)

print(f"   {len(results)} rows written")

# Add formulas
print("\n4. Adding formulas...")

last_row = len(results) + 2

# Total row
ws[f'A{last_row}'] = "TOTAL"
ws[f'E{last_row}'] = f'=SUM(E2:E{last_row-1})'
ws[f'F{last_row}'] = f'=SUM(F2:F{last_row-1})'

# Average row
ws[f'A{last_row+1}'] = "AVERAGE"
ws[f'E{last_row+1}'] = f'=AVERAGE(E2:E{last_row-1})'
ws[f'F{last_row+1}'] = f'=AVERAGE(F2:F{last_row-1})'

print("   Formulas added (SUM, AVERAGE)")

# Adjust column width
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 15

# Save file
print("\n5. Saving Excel file...")

wb.save('output/rfm_analysis.xlsx')

print("   File saved: output/rfm_analysis.xlsx")

# Create Segment Summary Sheet
print("\n6. Creating segment summary...")

# Get all customer data
cursor.execute('''
SELECT 
    c.customer_id,
    c.customer_name,
    SUM(o.order_amount) as total_spent
FROM customers c
LEFT JOIN orders o ON c.customer_id = o.customer_id
GROUP BY c.customer_id, c.customer_name
''')

all_customers = cursor.fetchall()

# Segment customers based on spending
segments = {'VIP': [], 'Loyal': [], 'Regular': [], 'At Risk': []}

for cust_id, name, total in all_customers:
    if total and total > 150000:
        segments['VIP'].append((cust_id, name, total))
    elif total and total > 100000:
        segments['Loyal'].append((cust_id, name, total))
    elif total and total > 50000:
        segments['Regular'].append((cust_id, name, total))
    else:
        segments['At Risk'].append((cust_id, name, total))

# Create segment summary sheet
ws2 = wb.create_sheet("Segment Summary")

ws2.append(["Segment", "Customer Count", "Total Revenue", "Average Order"])

for segment_name, customers in segments.items():
    if customers:
        count = len(customers)
        total_rev = sum(c[2] for c in customers if c[2])
        avg_order = total_rev / count if count > 0 else 0
        ws2.append([segment_name, count, total_rev, avg_order])

# Adjust column widths
ws2.column_dimensions['A'].width = 15
ws2.column_dimensions['B'].width = 18
ws2.column_dimensions['C'].width = 18
ws2.column_dimensions['D'].width = 18

# Save updated workbook
wb.save('output/rfm_analysis.xlsx')

print("   Segment summary added")

print("\n" + "=" * 80)
print("DAY 3 COMPLETE!")
print("=" * 80)

print("\nOutput File: output/rfm_analysis.xlsx")
print("Sheets: RFM Analysis, Segment Summary")

conn.close()